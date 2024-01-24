from tkinter import Tk, Label, Entry, Button, StringVar
from jira import JIRA
import time
import os
import openpyxl
from datetime import datetime
from ttkthemes import ThemedTk

desktop_path = os.path.join(os.path.expanduser("~"), "Documents")

# Jira login data
jira_server = "https://jira.bbraun.com"  # URL to your Jira server
jira_username = "Jira_PDT_01"  # Your username
jira_password = "Ji#bbmag#PDT#2023"  # Your password

# Connect to Jira
try:
    jira = JIRA(server=jira_server, basic_auth=(jira_username, jira_password))
except Exception as e:
    print(f'Error connecting to Jira: {e}')
    exit(1)

def get_input():
    start_date = start_entry.get()
    end_date = end_entry.get()
    ticket_number = ticket_entry.get()
    status = status_entry.get()
    priority = priority_entry.get()

    if start_date and end_date:
        jql = f'project = PDT AND created >= {start_date} AND created <= {end_date}'
    elif ticket_number:
        jql = f'project = PDT AND key = {ticket_number}'
    else:
        jql = 'project = PDT AND created >= -750d'

    if status:
        jql += f' AND status = "{status}"'
    if priority:
        jql += f' AND priority = "{priority}"'

    # Create or open Excel file
    filename = desktop_path + "\Dokumentation_Tickets.xlsx"
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(['Ticket number', 'Title', 'Description', 'Priority', 'Status','Created on', 'Created time', 'Resolved on', 'Resolved time','Reporter','Komponente'])

    issues = []
    start_at = 0
    max_results = 1000
    while True:
        new_tickets = jira.search_issues(jql, startAt=start_at, maxResults=max_results)
        if not new_tickets:
            break
        issues.extend(new_tickets)
        start_at += max_results

    worksheet.delete_rows(2, worksheet.max_row)

    for ticket in issues:
        issue = jira.issue(ticket.key)
        ticket_number = ticket.key
        title = issue.fields.summary
        description = issue.fields.description if issue.fields.description else ""
        priority = issue.fields.priority.name
        status = issue.fields.status.name
        Komponente = issue.fields.components[0].name if issue.fields.components else ""  # Added this line
        created_datetime = datetime.strptime(issue.fields.created, '%Y-%m-%dT%H:%M:%S.%f%z')
        created_on = created_datetime.strftime('%Y-%m-%d')
        created_time = created_datetime.strftime('%H:%M:%S')
        resolved_datetime = datetime.strptime(issue.fields.resolutiondate, '%Y-%m-%dT%H:%M:%S.%f%z') if issue.fields.resolutiondate else None
        resolved_on = resolved_datetime.strftime('%Y-%m-%d') if resolved_datetime else ""
        resolved_time = resolved_datetime.strftime('%H:%M:%S') if resolved_datetime else ""
        reporter = issue.fields.reporter.displayName if issue.fields.reporter else ""
        worksheet.append([ticket_number, title, description, priority, status, created_on, created_time, resolved_on, resolved_time, reporter, Komponente])  # Added 'Komponente' here
       # print(f'New ticket found: {ticket_number} - {title} - {priority}-{Komponente}')  # Output system message

    # Save Excel file
    workbook.save(filename)
    # Open Excel file
    os.startfile(filename)

root = ThemedTk(theme="arc")  # Wählen Sie ein Thema aus, das Ihnen gefällt
root.title('Jira Ticket Suche')

start_var = StringVar()
end_var = StringVar()
ticket_var = StringVar()
status_var = StringVar()
priority_var = StringVar()

Label(root, text='Start Datum(YYYY-MM-DD) oder leer lassen:').grid(row=0, column=0, sticky='w', padx=10, pady=10)
Entry(root, textvariable=start_var).grid(row=0, column=1, padx=10, pady=10)

Label(root, text='End Datum(YYYY-MM-DD) oder leer lassen:').grid(row=1, column=0, sticky='w', padx=10, pady=10)
Entry(root, textvariable=end_var).grid(row=1, column=1, padx=10, pady=10)

Label(root, text='Ticketnummer oder leer lassen:').grid(row=2, column=0, sticky='w', padx=10, pady=10)
Entry(root, textvariable=ticket_var).grid(row=2, column=1, padx=10, pady=10)

Label(root, text='Status oder leer lassen:').grid(row=3, column=0, sticky='w', padx=10, pady=10)
Entry(root, textvariable=status_var).grid(row=3, column=1, padx=10, pady=10)

Label(root, text='Priorität oder leer lassen:').grid(row=4, column=0, sticky='w', padx=10, pady=10)
Entry(root, textvariable=priority_var).grid(row=4, column=1, padx=10, pady=10)

Button(root, text='Bestätigen', command=get_input).grid(row=5, column=0, columnspan=5, padx=10, pady=10)

root.mainloop()